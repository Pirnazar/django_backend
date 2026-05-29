"""
Bulk admin actions for ItemAdmin.
"""
import io
import logging
from datetime import date

from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.utils.translation import gettext_lazy as _

from apps.common.choices import DeliveryStatus, ShipmentGroupStatus
from apps.shipments.models import ShipmentGroup
from .models import Item, ItemStatusHistory

_VALID_DELIVERY_STATUSES = {value for value, _label in DeliveryStatus.choices}

logger = logging.getLogger(__name__)

# State machine copied from apps/items/views.py
ALLOWED_TRANSITIONS = {
    DeliveryStatus.CREATED:              [DeliveryStatus.AT_CHINA_WAREHOUSE, DeliveryStatus.CANCELLED],
    DeliveryStatus.AT_CHINA_WAREHOUSE:   [DeliveryStatus.MEASURED],
    DeliveryStatus.MEASURED:             [DeliveryStatus.PHOTOGRAPHED, DeliveryStatus.LABELED],
    DeliveryStatus.PHOTOGRAPHED:         [DeliveryStatus.LABELED],
    DeliveryStatus.LABELED:              [DeliveryStatus.PACKED],
    DeliveryStatus.PACKED:               [DeliveryStatus.GROUPED],
    DeliveryStatus.GROUPED:              [DeliveryStatus.SENT_TO_URUMQI],
    DeliveryStatus.SENT_TO_URUMQI:       [DeliveryStatus.ARRIVED_URUMQI],
    DeliveryStatus.ARRIVED_URUMQI:       [DeliveryStatus.SENT_TO_TURKMENISTAN, DeliveryStatus.OUT_FOR_DELIVERY],
    DeliveryStatus.SENT_TO_TURKMENISTAN: [DeliveryStatus.ARRIVED_TURKMENISTAN],
    DeliveryStatus.ARRIVED_TURKMENISTAN: [DeliveryStatus.OUT_FOR_DELIVERY],
    DeliveryStatus.OUT_FOR_DELIVERY:     [DeliveryStatus.DELIVERED],
    DeliveryStatus.DELIVERED:            [],
    DeliveryStatus.CANCELLED:            [],
}


def action_add_to_group(modeladmin, request, queryset):
    """Добавить выбранные грузы в партию."""
    items = list(queryset.select_related('destination', 'warehouse'))

    if 'apply' in request.POST:
        group_id = request.POST.get('group_id')
        try:
            group = ShipmentGroup.objects.get(pk=group_id)
        except ShipmentGroup.DoesNotExist:
            modeladmin.message_user(request, 'Партия не найдена.', messages.ERROR)
            return HttpResponseRedirect(request.get_full_path())

        # Validate all items share same destination & warehouse as group
        bad = [
            item.item_code for item in items
            if item.destination_id != group.destination_id or item.warehouse_id != group.warehouse_id
        ]
        if bad:
            modeladmin.message_user(
                request,
                f'Грузы {", ".join(bad)} не совместимы с партией (другой склад/направление).',
                messages.ERROR,
            )
            return HttpResponseRedirect(request.get_full_path())

        count = 0
        for item in items:
            item.shipment_group = group
            item.updated_by = request.user
            item.save(update_fields=['shipment_group', 'updated_by'])
            count += 1

        modeladmin.message_user(
            request,
            f'Добавлено {count} груз(ов) в партию {group.group_code}.',
            messages.SUCCESS,
        )
        return HttpResponseRedirect(request.get_full_path())

    # GET-phase: detect common destination/warehouse from first item
    first = items[0] if items else None
    compatible_groups = ShipmentGroup.objects.exclude(
        status__in=[ShipmentGroupStatus.COMPLETED, ShipmentGroupStatus.CANCELLED]
    )
    if first:
        compatible_groups = compatible_groups.filter(
            destination=first.destination,
            warehouse=first.warehouse,
        )

    return TemplateResponse(
        request,
        'admin/items/bulk_add_to_group.html',
        {
            'title': 'Добавить грузы в партию',
            'items': items,
            'groups': compatible_groups.order_by('-created_at')[:50],
            'action_checkbox_name': '_selected_action',
        },
    )


action_add_to_group.short_description = _('Добавить выбранные грузы в партию')


def action_change_status(modeladmin, request, queryset):
    """Изменить статус выбранных грузов."""
    items = list(queryset.select_related('client'))

    if 'apply' in request.POST:
        new_status = request.POST.get('new_status', '')
        comment = request.POST.get('comment', '')

        if not new_status:
            modeladmin.message_user(request, 'Не выбран новый статус.', messages.ERROR)
            return HttpResponseRedirect(request.get_full_path())

        if new_status not in _VALID_DELIVERY_STATUSES:
            modeladmin.message_user(request, 'Недопустимый статус доставки.', messages.ERROR)
            return HttpResponseRedirect(request.get_full_path())

        new_label = DeliveryStatus(new_status).label
        ok, errors = [], []
        for item in items:
            allowed = ALLOWED_TRANSITIONS.get(item.delivery_status, [])
            if new_status not in allowed:
                errors.append(
                    f'{item.item_code}: нельзя перевести из '
                    f'"{item.get_delivery_status_display()}" → "{new_label}"'
                )
                continue
            old_status = item.delivery_status
            item.delivery_status = new_status
            item.updated_by = request.user
            item.save(update_fields=['delivery_status', 'updated_by'])
            ItemStatusHistory.objects.create(
                item=item,
                old_status=old_status,
                new_status=new_status,
                comment=comment,
                changed_by=request.user,
            )
            ok.append(item.item_code)

        if ok:
            modeladmin.message_user(
                request,
                f'Статус обновлён для {len(ok)} груз(ов).',
                messages.SUCCESS,
            )
        for err in errors:
            modeladmin.message_user(request, err, messages.WARNING)

        return HttpResponseRedirect(request.get_full_path())

    status_choices = [
        (v, label) for v, label in DeliveryStatus.choices
        if v not in (DeliveryStatus.CREATED,)
    ]
    return TemplateResponse(
        request,
        'admin/items/bulk_change_status.html',
        {
            'title': 'Изменить статус грузов',
            'items': items,
            'status_choices': status_choices,
            'action_checkbox_name': '_selected_action',
        },
    )


action_change_status.short_description = _('Изменить статус выбранных грузов')


def action_export_excel(modeladmin, request, queryset):
    """Экспортировать выбранные грузы в Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        modeladmin.message_user(
            request, 'Библиотека openpyxl не установлена.', messages.ERROR
        )
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Грузы'

    headers = [
        'Код груза', 'Код клиента', 'Клиент', 'Телефон',
        'Направление', 'Склад', 'Партия',
        'Вес (кг)', 'Объём (м³)', 'Сумма',
        'Оплата', 'Статус доставки', 'Дата создания',
    ]
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    qs = queryset.select_related('client', 'destination', 'warehouse', 'shipment_group')
    for row_idx, item in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1,  value=item.item_code)
        ws.cell(row=row_idx, column=2,  value=item.client.client_code if item.client else '')
        ws.cell(row=row_idx, column=3,  value=item.client.full_name if item.client else '')
        ws.cell(row=row_idx, column=4,  value=item.client.phone_number if item.client else '')
        ws.cell(row=row_idx, column=5,  value=item.destination.code if item.destination else '')
        ws.cell(row=row_idx, column=6,  value=item.warehouse.code if item.warehouse else '')
        ws.cell(row=row_idx, column=7,  value=item.shipment_group.group_code if item.shipment_group else '')
        ws.cell(row=row_idx, column=8,  value=float(item.weight_kg))
        ws.cell(row=row_idx, column=9,  value=float(item.volume_m3))
        ws.cell(row=row_idx, column=10, value=float(item.total_price))
        ws.cell(row=row_idx, column=11, value=item.get_payment_status_display())
        ws.cell(row=row_idx, column=12, value=item.get_delivery_status_display())
        ws.cell(row=row_idx, column=13, value=item.created_at.strftime('%d.%m.%Y %H:%M') if item.created_at else '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'items_export_{date.today():%Y%m%d}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


action_export_excel.short_description = _('Экспортировать выбранные грузы в Excel')


def action_label_preview(modeladmin, request, queryset):
    """Печать этикеток выбранных грузов."""
    items = queryset.select_related('client', 'destination', 'warehouse')
    return TemplateResponse(
        request,
        'admin/items/bulk_label_preview.html',
        {
            'title': 'Предпросмотр этикеток',
            'items': items,
        },
    )


action_label_preview.short_description = _('Печать этикеток выбранных грузов')
